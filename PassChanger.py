import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) 

import sys
import os
import telnetlib
import paramiko
import openpyxl
import logging
import datetime

USER_LOGIN = "u0_a292"
TIMEOUT = 3

IP_COLUMN_NAME = "ip"
PORT_COLUMN_NAME = "port"
PASSWORD_COLUMN_NAME = "password"
SHEET_NAME = "PassSheet"

WHATIF_MODE = None
USE_PORT = False

logInfo: logging.Logger
logError: logging.Logger

def setup_loggers():
    """Настройка логирования"""
    global logInfo, logError

    current_day = datetime.datetime.now().strftime("%Y_%m_%d")
    log_info_file = f"Logs/Info/{current_day}_info.log"
    log_error_file = f"Logs/Errors/{current_day}_error.log"

    os.makedirs(os.path.dirname(log_info_file), exist_ok=True)
    os.makedirs(os.path.dirname(log_error_file), exist_ok=True)
    
    logInfo = get_logger(log_info_file, logging.INFO)
    logError = get_logger(log_error_file, logging.ERROR)


def get_logger(log_file: str, level: int = logging.INFO):
    """Создаёт новый логер для записи в разные файлы в зависимости от уровня сообщения"""
    handler = logging.FileHandler(log_file, encoding="utf-8")
    handler.setFormatter(logging.Formatter("[%(asctime)s] %(message)s", "%Y-%m-%d %H:%M:%S"))

    logger = logging.getLogger(log_file)
    logger.setLevel(level)
    logger.addHandler(handler)

    return logger

def log_format(ip, msg, exception) -> str:
    """Форматирование для вывода в три колонки"""
    return f"{ip.ljust(15)} {msg.ljust(45)} {str(exception)}"


def connect_telnet(username, password, new_password, ip, port = 23) -> bool:
    """Подключение Telnet для смены пароля"""
    try:
        if (USE_PORT):
            tn = telnetlib.Telnet(ip, port, timeout=TIMEOUT)
        else:
            tn = telnetlib.Telnet(ip, timeout=TIMEOUT)
            
        tn.read_until(b"login: ", timeout=TIMEOUT)
        tn.write(username.encode("utf-8") + b"\n")
        tn.read_until(b"Password: ")
        tn.write(password.encode("utf-8") + b"\n")
        
        output = tn.read_until(b"# ", timeout=TIMEOUT)
        
        # Обработка ошибки авторизации
        if b"Login incorrect" in output:
            logError.error(log_format(ip, "Telnet: Неверное имя пользователя или пароль", ""))
            return False
        
        if WHATIF_MODE:
            logInfo.info(f"{ip.ljust(15)} WHATIF: Пароль будет изменён на: {new_password}")
        else:
            tn.write(b"password\n")
            tn.read_until(b"New password: ")
            tn.write(new_password.encode("utf-8") + b"\n")
            tn.read_until(b"Retype new password: ")
            tn.write(new_password.encode("utf-8") + b"\n")

            logInfo.info(f"{ip.ljust(15)} - Пароль изменён на: {new_password}")

        tn.write(b"exit\n")
        tn.read_all()
        tn.close()
        return True
    except Exception as e:
        logError.error(log_format(ip, "Telnet: Не удалось подключиться к устройству", e))
        return False

def connect_ssh(username, password, new_password, ip, port = 22) -> bool:
    """Подключение SSH для смены пароля"""
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        if(USE_PORT):
            client.connect(ip, port, username=username, password=password, timeout=TIMEOUT)
        else:
            client.connect(ip, username=username, password=password, timeout=TIMEOUT)
        
        if WHATIF_MODE:
            logInfo.info(f"{ip.ljust(15)} WHATIF: Пароль будет изменён на: {new_password}")
        else:
            stdin, stdout, stderr = client.exec_command(f"passwd")
            stdin.write(password + "\n")
            stdin.write(new_password + "\n")
            stdin.write(new_password + "\n")
            stdin.flush()

            logInfo.info(f"{ip.ljust(15)} Пароль изменён на: {new_password}")

        client.close()
        return True
    except paramiko.AuthenticationException as e:
        logError.error(log_format(ip, "SSH: Ошибка авторизации", e))
        return False
    except Exception as e:
        logError.error(log_format(ip, "SSH: Не удалось подключиться к устройству", e))
        return False

def connect(username, password, new_password, ip, port = None) -> bool:
    """Подключение для смены пароля"""
    if USE_PORT and port is not None:
        return connect_telnet(username, password, new_password, ip, port) or connect_ssh(username, password, new_password, ip, port)
    else:
        return connect_telnet(username, password, new_password, ip) or connect_ssh(username, password, new_password, ip)

def create_results_file(results) -> None:
    """Записывает полученные результаты в xlsx файл"""
    output_file = "result.xlsx"
    sheet_name = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    # Если файл уже существует, загрузить его
    if os.path.exists(output_file):
        workbook = openpyxl.load_workbook(output_file)
    else:
        workbook = openpyxl.Workbook()

    worksheet = workbook.create_sheet(title=sheet_name)

    # Запись заголовков только если создан новый лист
    if sheet_name == workbook.sheetnames[-1]:
        headers = ["ip", "old_password", "new_password", "changed"]
        for col_num, header in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)  # type: ignore
            worksheet[column_letter + "1"] = header

    # Запись данных
    for row_num, result in enumerate(results, start=2):
        worksheet["A" + str(row_num)] = result["ip"]
        worksheet["B" + str(row_num)] = result["old_password"]
        worksheet["C" + str(row_num)] = result["new_password"]
        worksheet["D" + str(row_num)] = result["changed"]

    workbook.save(output_file)
    logInfo.info(f"Результаты сохранены в файле: {output_file}")

def ask_yes_no_question(question):
    """Подтверждение выполнения"""
    valid = {"yes": True, "y": True, "ye": True, "no": False, "n": False}
    while True:
        sys.stdout.write(question + " [y/n] ")
        choice = input().lower()
        if choice in valid:
            return valid[choice]
        else:
            sys.stdout.write("Пожалуйста, ответьте 'yes' или 'no'\n")

def main() -> None:
    """Главный метод"""
    global WHATIF_MODE
    setup_loggers()

    # Чтение аргументов и установка режима WHATIF
    logInfo.info("Скрипт запущен")
    args = sys.argv[1:]
    WHATIF_MODE = "-whatif" in args

    if WHATIF_MODE:
        args.remove("-whatif")

    if len(args) < 2:
        print("Usage: PassChanger.py [-whatif] <excel_file> <passwords_file>")
        return

    excel_file = args[0]
    password_file = args[1]

    # Парсинг эксель файла
    logInfo.info(f"Получение данных из файла: {excel_file}")
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[SHEET_NAME]
    headers = [cell.value for cell in sheet[1]]
    ip_column_index = headers.index(IP_COLUMN_NAME)
    password_column_index = headers.index(PASSWORD_COLUMN_NAME)
    port_column_index = headers.index(PORT_COLUMN_NAME) if USE_PORT else -1

    results = []

    if not WHATIF_MODE:
        if not ask_yes_no_question(f"Вы действительно хотите изменить {sheet.max_row-1} паролей?"):
            sys.exit()

    # Чтение паролей из файла
    with open(password_file, "r", encoding="UTF-8") as file:
        passwords = file.readlines()
        passwords = [password.strip() for password in passwords]
        password_count = len(passwords)

    # Итеративная обработка записей с попыткой подключения через Telnet или SSH
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        ip = row[ip_column_index]
        port = row[port_column_index]
        password = row[password_column_index]

        # Если паролей меньше чем устройств, то начинает сначала
        password_index = (index - 1) % password_count
        new_password = passwords[password_index]

        print(f"{index}/{sheet.max_row-1} Подключение к устройству {ip}")
        password_changed = connect(USER_LOGIN, password, new_password, ip, port) 

        result = {
            "ip": ip,
            "old_password": password,
            "new_password": new_password,
            "changed": password_changed,
        }
        results.append(result)

    create_results_file(results)
    logInfo.info("Скрипт выполнен")

if __name__ == "__main__":
    main()
